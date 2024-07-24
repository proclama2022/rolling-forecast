import io
import os
from anthropic import Anthropic
import fitz  # PyMuPDF per la gestione dei PDF
from PIL import Image
import pytesseract
import docx2txt
import openpyxl

def generate_forecast_with_claude(forecast_type, historical_data, info, assumptions, forecast_period):
    prompt = f"""
    Genera un rolling forecast per i prossimi {forecast_period} mesi per {forecast_type} basato sui seguenti dati:

    Dati storici:
    {historical_data}

    Informazioni:
    {info}

    Assunzioni di crescita:
    {assumptions}

    Fornisci una previsione dettagliata in formato CSV con le seguenti colonne (separati da virgole):
    Anno,Mese,{forecast_type.capitalize()}_Totali,Unità,Prezzo,Vendite_Totali,Percentuale_Totale

    Aggiungi una riga per i totali annuali alla fine di ogni anno.
    Assicurati che il CSV abbia esattamente 7 colonne per ogni riga, compresi i totali annuali.
    """

    response = anthropic.messages.create(
        model="claude-3-sonnet-20240229",
        max_tokens=4000,
        messages=[
            {"role": "user", "content": prompt}
        ]
    )

    return response.content[0].text

def analyze_forecast_with_ai(sales_forecast, cost_forecast, forecast_period):
    combined_data = f"""
    Previsione Vendite:
    {sales_forecast.to_string()}

    Previsione Costi:
    {cost_forecast.to_string()}
    """

    prompt = f"""
    Analizza le seguenti previsioni di vendite e costi per i prossimi {forecast_period} mesi:

    {combined_data}

    Fornisci un'analisi dettagliata che includa:
    1. Tendenze generali nelle vendite e nei costi
    2. Periodi di picco e di bassa per vendite e costi
    3. Analisi della redditività nel tempo
    4. Potenziali rischi o opportunità identificati
    5. Raccomandazioni per migliorare la performance finanziaria

    Presenta l'analisi in un formato chiaro e strutturato.
    """

    response = anthropic.messages.create(
        model="claude-3-sonnet-20240229",
        max_tokens=4000,
        messages=[
            {"role": "user", "content": prompt}
        ]
    )

    return response.content[0].text

def extract_data_from_balance(uploaded_file):
    file_extension = os.path.splitext(uploaded_file.name)[1].lower()
    text = ""

    if file_extension == '.pdf':
        pdf_document = fitz.open(stream=uploaded_file.getvalue(), filetype="pdf")
        for page_num in range(len(pdf_document)):
            page = pdf_document[page_num]
            text += page.get_text()  # Estrae tutto il testo

    elif file_extension == '.docx':
        text = docx2txt.process(io.BytesIO(uploaded_file.getvalue()))
    elif file_extension in ['.xlsx', '.xls']:
        wb = openpyxl.load_workbook(io.BytesIO(uploaded_file.getvalue()))
        for sheet in wb.sheetnames:
            for row in wb[sheet].iter_rows(values_only=True):
                text += " ".join(str(cell) for cell in row if cell is not None) + "\n"
    elif file_extension == '.txt':
        text = uploaded_file.getvalue().decode('utf-8')
    else:
        raise ValueError(f"Formato file non supportato: {file_extension}")

    return text

def parse_extracted_data(text, extraction_criteria):
    prompt = f"""
    Estrai i seguenti dati dal testo:
    {extraction_criteria}

    Testo:
    {text}
    """

    response = anthropic.messages.create(
        model="claude-3-sonnet-20240229",
        max_tokens=4000,
        messages=[
            {"role": "user", "content": prompt}
        ]
    )

    # Supponiamo che Claude restituisca dati in formato strutturato
    structured_data = response.content[0].text

    # Potrebbe restituire dati in formato JSON o simile, quindi possiamo convertirli in un dizionario
    try:
        extracted_data = eval(structured_data)  # Assicurati che Claude restituisca un formato sicuro
    except Exception as e:
        raise ValueError("Errore durante il parsing dei dati estratti: " + str(e))

    return extracted_data
